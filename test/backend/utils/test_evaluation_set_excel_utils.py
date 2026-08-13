"""Unit tests for ``backend.utils.evaluation_set_excel_utils``.

Tests cover Excel template generation and case parsing for both .xlsx
and .xls (legacy ``xlrd`` path) formats, including bilingual header
resolution (Chinese / English), required-column validation, row-level
error reporting, and empty-file handling.

The parser supports exactly four columns:

  - ``session_id`` / ``会话ID``
  - ``request_id`` / ``请求顺序`` / ``turn_order``
  - ``query`` / ``问题``                       (required)
  - ``answer`` / ``答案`` / ``reference_output``
"""

import importlib
import io
import os
import sys
import types

import pytest


@pytest.fixture(autouse=True)
def _ensure_real_openpyxl_loaded_for_module_imports():
    """Reload the production module with a fresh ``openpyxl`` before each test.

    Why we don't use ``monkeypatch.setattr(sys.modules, ...)``: the
    production parser binds ``openpyxl.load_workbook`` at module-import
    time.  A sibling fixture that replaces ``sys.modules["openpyxl"]``
    between two test runs of *this* file would leave the parser's bound
    reference pointing at the mock.

    Reloading the parser inside this fixture (which is autouse=True and
    therefore runs before any service-level fixtures) re-binds
    ``openpyxl.load_workbook`` to the freshly-loaded real module.  No
    module-level mutation persists past the test.
    """
    sys.modules.pop("openpyxl", None)
    sys.modules.pop("openpyxl.styles", None)
    importlib.import_module("openpyxl")
    importlib.import_module("openpyxl.styles")
    parser = sys.modules.get("backend.utils.evaluation_set_excel_utils")
    if parser is not None:
        importlib.reload(parser)
    yield

# Force real ``openpyxl`` to win over any MagicMock a sibling test module may
# have installed into ``sys.modules`` earlier in the session.  We must do this
# *before* the module under test is imported, otherwise the bound names
# inside ``evaluation_set_excel_utils`` will reference the MagicMock and
# fail when called.
try:
    _real_openpyxl = importlib.import_module("openpyxl")
    if not callable(getattr(_real_openpyxl, "Workbook", None)):
        raise ImportError("openpyxl.Workbook is not callable")
except Exception:
    # Fallback: try to import it via the OpenPyXL wheels if available.
    sys.modules.pop("openpyxl", None)
    sys.modules.pop("openpyxl.styles", None)
    _real_openpyxl = importlib.import_module("openpyxl")

sys.modules["openpyxl"] = _real_openpyxl
sys.modules["openpyxl.styles"] = importlib.import_module("openpyxl.styles")

# NOTE: ``xlrd`` is optionally stubbed in conftest.py because the production
# code imports it at module load.  For these tests we need a richer fake that
# actually returns a sheet we can drive — the conftest stub is a bare
# ``MagicMock`` which works for the .xlsx path (which doesn't use ``xlrd``)
# but would not survive a .xls parse.  We override the conftest stub here
# before importing the module under test so both paths get a working fake.
class _StubSheet:
    def __init__(self, rows):
        # rows: list of lists of cell values (header row first)
        self.nrows = len(rows)
        self._rows = rows

    def row_values(self, rowx):
        return self._rows[rowx] if rowx < len(self._rows) else []

    def cell_value(self, rowx, colx):
        row = self._rows[rowx] if rowx < len(self._rows) else []
        return row[colx] if colx < len(row) else None


class _StubBook:
    def __init__(self, rows):
        self._sheet = _StubSheet(rows)

    def sheet_by_index(self, idx):
        return self._sheet


class _StubXlrd:
    @staticmethod
    def open_workbook(file_contents=b""):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(file_contents), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        return _StubBook(rows)


# Replace whatever conftest set with our richer fake.  This must run before
# the module under test is (re-)loaded so that the .xls path uses our fake.
_xlrd_stub = types.ModuleType("xlrd")
_xlrd_stub.open_workbook = _StubXlrd.open_workbook
sys.modules["xlrd"] = _xlrd_stub

# Ensure backend is on the path.
_BACKEND = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "backend"))
if _BACKEND not in sys.path:
    sys.path.insert(0, _BACKEND)

import importlib as _importlib

# Always force real ``openpyxl`` into sys.modules before reloading the
# production module — sibling fixtures may have left a MagicMock stub there.
for _name in ("openpyxl", "openpyxl.styles", "openpyxl.workbook"):
    sys.modules.pop(_name, None)
try:
    import openpyxl
    sys.modules["openpyxl"] = openpyxl
except Exception:
    pass

_existing_module = sys.modules.get("backend.utils.evaluation_set_excel_utils")
if _existing_module is None:
    _module = _importlib.import_module("backend.utils.evaluation_set_excel_utils")
elif getattr(_existing_module, "xlrd", None) is not _xlrd_stub:
    _module = _importlib.reload(_existing_module)
else:
    _module = _existing_module

# Bind public symbols used by tests.
REQUIRED_FIELDS = _module.REQUIRED_FIELDS
ALL_FIELDS = _module.ALL_FIELDS
HEADER_ALIASES = _module.HEADER_ALIASES
_normalize_header = _module._normalize_header
_canonical_header = _module._canonical_header
build_evaluation_set_excel_template_bytes = _module.build_evaluation_set_excel_template_bytes
build_evaluation_set_export_bytes = _module.build_evaluation_set_export_bytes
parse_evaluation_cases_from_excel = _module.parse_evaluation_cases_from_excel


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_xlsx_bytes(header_row, data_rows):
    """Build a minimal in-memory .xlsx file from row data.

    header_row: list of column headers (str)
    data_rows : list of lists, each row's cell values

    Always uses the real ``openpyxl.Workbook`` even if a sibling test file has
    stubbed ``sys.modules["openpyxl"]`` for its own purposes.
    """
    sys.modules.pop("openpyxl", None)
    sys.modules.pop("openpyxl.styles", None)
    sys.modules.pop("openpyxl.workbook", None)
    sys.modules.pop("openpyxl.workbook.workbook", None)
    Workbook = importlib.import_module("openpyxl").Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(header_row)
    for row in data_rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# _normalize_header
# ---------------------------------------------------------------------------

class TestNormalizeHeader:
    def test_none_returns_empty_string(self):
        assert _normalize_header(None) == ""

    def test_strips_and_lowercases(self):
        # Lowercasing is a no-op for pure Chinese strings.
        assert _normalize_header("  ANSWER  ") == "answer"
        assert _normalize_header("问题") == "问题"

    def test_trailing_star_preserved(self):
        # _normalize_header itself does NOT strip the *; _canonical_header
        # is responsible for tolerating the required-marker.
        assert _normalize_header("query*") == "query*"


# ---------------------------------------------------------------------------
# _canonical_header
# ---------------------------------------------------------------------------

class TestCanonicalHeader:
    def test_none_returns_none(self):
        assert _canonical_header(None) is None

    def test_empty_string_returns_none(self):
        assert _canonical_header("") is None
        assert _canonical_header("   ") is None

    def test_english_headers(self):
        assert _canonical_header("session_id") == "session_id"
        assert _canonical_header("request_id") == "request_id"
        assert _canonical_header("query") == "query"
        assert _canonical_header("answer") == "answer"

    def test_english_headers_case_insensitive(self):
        assert _canonical_header("SESSION_ID") == "session_id"
        assert _canonical_header("Query") == "query"
        assert _canonical_header("Answer") == "answer"

    def test_chinese_headers(self):
        assert _canonical_header("会话ID") == "session_id"
        assert _canonical_header("请求顺序") == "request_id"
        assert _canonical_header("问题") == "query"
        assert _canonical_header("答案") == "answer"

    def test_reference_output_alias(self):
        assert _canonical_header("reference_output") == "answer"
        assert _canonical_header("expected_output") == "answer"

    def test_turn_order_alias(self):
        assert _canonical_header("turn_order") == "request_id"
        assert _canonical_header("turn") == "request_id"

    def test_trailing_star_is_tolerated(self):
        assert _canonical_header("query*") == "query"
        assert _canonical_header("问题*") == "query"

    def test_unknown_header_returns_none(self):
        assert _canonical_header("foo") is None
        assert _canonical_header("custom_variables") is None
        assert _canonical_header("case_id") is None


# ---------------------------------------------------------------------------
# build_evaluation_set_excel_template_bytes
# ---------------------------------------------------------------------------

class TestBuildTemplateBytes:
    def test_returns_bytes(self):
        result = build_evaluation_set_excel_template_bytes()
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_is_valid_xlsx(self):
        """The generated bytes can be loaded back by openpyxl."""
        from openpyxl import load_workbook

        result = build_evaluation_set_excel_template_bytes()
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.title == "evaluation_cases"

    def test_zh_template_has_four_chinese_headers(self):
        """zh template row 2 should be the four Chinese headers."""
        from openpyxl import load_workbook

        result = build_evaluation_set_excel_template_bytes(language="zh")
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        # Row 1 — instruction descriptions (Chinese)
        row1 = [cell.value for cell in ws[1]]
        assert any("对话" in str(v) for v in row1)
        # Row 2 — four Chinese column headers
        row2 = [cell.value for cell in ws[2]]
        assert row2 == ["会话ID", "请求顺序", "问题", "答案"]

    def test_zh_template_instructions_are_detailed(self):
        """Row 1 instructions explain required/optional and multi-turn rules."""
        from openpyxl import load_workbook

        result = build_evaluation_set_excel_template_bytes(language="zh")
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        cells = [str(cell.value) for cell in ws[1]]
        assert any("必填" in v for v in cells)
        assert any("选填" in v for v in cells)
        assert any("同一会话" in v for v in cells)
        assert any("多轮对话" in v for v in cells)

    def test_instruction_row_uses_wrap_text_and_height(self):
        """The instruction row wraps text and has a fixed height."""
        from openpyxl import load_workbook

        result = build_evaluation_set_excel_template_bytes(language="zh")
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        assert cell.alignment.wrap_text is True
        assert ws.row_dimensions[1].height == 64

    def test_en_template_has_four_english_headers(self):
        """en template row 2 should be the four English headers."""
        from openpyxl import load_workbook

        result = build_evaluation_set_excel_template_bytes(language="en")
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        row1 = [cell.value for cell in ws[1]]
        # Row 1 contains English instruction text (no Chinese).
        assert any("conversation" in str(v) for v in row1)
        row2 = [cell.value for cell in ws[2]]
        assert row2 == ["session_id", "request_id", "query", "answer"]

    def test_default_language_is_zh(self):
        from openpyxl import load_workbook

        result = build_evaluation_set_excel_template_bytes()
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        row2 = [cell.value for cell in ws[2]]
        assert row2 == ["会话ID", "请求顺序", "问题", "答案"]

    def test_example_data_present(self):
        """Row 3+ should contain example data."""
        from openpyxl import load_workbook

        result = build_evaluation_set_excel_template_bytes()
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        # Row 3 should be an example data row (not header, not empty)
        row3 = [cell.value for cell in ws[3]]
        assert row3[2] is not None  # query column should have data

    def test_unknown_language_falls_back_to_zh(self):
        from openpyxl import load_workbook

        result = build_evaluation_set_excel_template_bytes(language="fr")
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        row2 = [cell.value for cell in ws[2]]
        assert row2 == ["会话ID", "请求顺序", "问题", "答案"]


# ---------------------------------------------------------------------------
# build_evaluation_set_export_bytes
# ---------------------------------------------------------------------------

class TestBuildExportBytes:
    def test_returns_bytes(self):
        result = build_evaluation_set_export_bytes([])
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_export_uses_zh_headers(self):
        """Export always uses the zh 4-column header layout."""
        from openpyxl import load_workbook

        cases = [
            {
                "inputs": {"query": "q1", "session_id": "s1", "request_id": "1"},
                "label": {"answer": "a1"},
            },
        ]
        result = build_evaluation_set_export_bytes(cases)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        row2 = [cell.value for cell in ws[2]]
        assert row2 == ["会话ID", "请求顺序", "问题", "答案"]
        row3 = [cell.value for cell in ws[3]]
        assert row3[0] == "s1"
        assert row3[1] == "1"
        assert row3[2] == "q1"
        assert row3[3] == "a1"

    def test_turn_order_fallback(self):
        """When inputs use 'turn_order' instead of 'request_id'."""
        from openpyxl import load_workbook

        cases = [
            {
                "inputs": {"query": "q1", "turn_order": "3"},
                "label": {"answer": "a1"},
            },
        ]
        result = build_evaluation_set_export_bytes(cases)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        row3 = [cell.value for cell in ws[3]]
        assert row3[1] == "3"

    def test_session_id_from_top_level(self):
        """When session_id is at case top-level, not in inputs."""
        from openpyxl import load_workbook

        cases = [
            {
                "inputs": {"query": "q1"},
                "label": {"answer": "a1"},
                "session_id": "top_level_s",
            },
        ]
        result = build_evaluation_set_export_bytes(cases)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        row3 = [cell.value for cell in ws[3]]
        assert row3[0] == "top_level_s"

    def test_export_then_import_round_trip(self):
        """An exported file should be parseable by the import parser."""
        cases = [
            {
                "inputs": {
                    "query": "round trip query",
                    "session_id": "rt1",
                    "request_id": "1",
                },
                "label": {"answer": "round trip answer"},
            },
        ]
        exported = build_evaluation_set_export_bytes(cases)
        parsed = parse_evaluation_cases_from_excel("rt.xlsx", exported)
        assert len(parsed) == 1
        assert parsed[0]["inputs"]["query"] == "round trip query"
        assert parsed[0]["inputs"]["session_id"] == "rt1"
        assert parsed[0]["label"]["answer"] == "round trip answer"


# ---------------------------------------------------------------------------
# parse_evaluation_cases_from_excel — shared behaviour for .xlsx and .xls
# ---------------------------------------------------------------------------

class TestParseShared:
    """Tests that apply to both .xlsx and .xls paths."""

    def test_unknown_file_extension_raises(self):
        raw = _make_xlsx_bytes(["query", "answer"], [["q", "a"]])
        with pytest.raises(ValueError, match="Unsupported file type"):
            parse_evaluation_cases_from_excel("cases.csv", raw)

    def test_header_normalization_trims_and_lowercases(self):
        raw = _make_xlsx_bytes(["  QUERY  ", " ANSWER "], [["q1", "a1"]])
        cases = parse_evaluation_cases_from_excel("test.xlsx", raw)
        assert len(cases) == 1
        assert cases[0]["inputs"]["query"] == "q1"

    def test_required_column_missing(self):
        raw = _make_xlsx_bytes(["session_id"], [["s1"]])
        with pytest.raises(ValueError, match="Missing required column"):
            parse_evaluation_cases_from_excel("test.xlsx", raw)

    def test_order_no_is_sequential(self):
        raw = _make_xlsx_bytes(
            ["query", "answer"],
            [["q1", "a1"], ["q2", "a2"], ["q3", "a3"]],
        )
        cases = parse_evaluation_cases_from_excel("test.xlsx", raw)
        assert [c["order_no"] for c in cases] == [0, 1, 2]

    def test_strips_whitespace_from_cell_values(self):
        raw = _make_xlsx_bytes(
            ["query", "answer"],
            [["  trimmed  ", "  spaced  "]],
        )
        cases = parse_evaluation_cases_from_excel("test.xlsx", raw)
        assert cases[0]["inputs"]["query"] == "trimmed"
        assert cases[0]["label"]["answer"] == "spaced"

    def test_trailing_star_on_header_is_tolerated(self):
        raw = _make_xlsx_bytes(["query*", "answer*"], [["q1", "a1"]])
        cases = parse_evaluation_cases_from_excel("test.xlsx", raw)
        assert len(cases) == 1
        assert cases[0]["inputs"]["query"] == "q1"

    def test_empty_cells_in_optional_column_ignored(self):
        raw = _make_xlsx_bytes(
            ["session_id", "query", "answer"],
            [["", "q1", "a1"]],
        )
        cases = parse_evaluation_cases_from_excel("test.xlsx", raw)
        assert cases[0].get("session_id") is None

    # ── Bilingual header parsing ────────────────────────────────────

    def test_parse_chinese_headers(self):
        raw = _make_xlsx_bytes(
            ["会话ID", "请求顺序", "问题", "答案"],
            [["s1", "1", "你好", "世界"]],
        )
        cases = parse_evaluation_cases_from_excel("test.xlsx", raw)
        assert len(cases) == 1
        assert cases[0]["inputs"]["query"] == "你好"
        assert cases[0]["inputs"]["session_id"] == "s1"
        assert cases[0]["inputs"]["request_id"] == "1"
        assert cases[0]["label"]["answer"] == "世界"
        assert cases[0]["session_id"] == "s1"
        assert cases[0]["turn_order"] == "1"

    def test_parse_english_headers(self):
        raw = _make_xlsx_bytes(
            ["session_id", "request_id", "query", "answer"],
            [["s1", "1", "hello", "world"]],
        )
        cases = parse_evaluation_cases_from_excel("test.xlsx", raw)
        assert len(cases) == 1
        assert cases[0]["inputs"]["query"] == "hello"
        assert cases[0]["inputs"]["session_id"] == "s1"
        assert cases[0]["label"]["answer"] == "world"

    def test_parse_mixed_chinese_and_english_headers(self):
        """Mixing zh and en headers in the same file should still work."""
        raw = _make_xlsx_bytes(
            ["session_id", "请求顺序", "query", "答案"],
            [["s1", "1", "q", "a"]],
        )
        cases = parse_evaluation_cases_from_excel("test.xlsx", raw)
        assert cases[0]["inputs"]["query"] == "q"
        assert cases[0]["inputs"]["session_id"] == "s1"
        assert cases[0]["inputs"]["request_id"] == "1"
        assert cases[0]["label"]["answer"] == "a"

    def test_turn_order_alias(self):
        raw = _make_xlsx_bytes(
            ["query", "turn_order"],
            [["q1", "5"]],
        )
        cases = parse_evaluation_cases_from_excel("test.xlsx", raw)
        assert cases[0]["inputs"]["request_id"] == "5"
        assert cases[0]["turn_order"] == "5"

    def test_reference_output_maps_to_label_answer(self):
        raw = _make_xlsx_bytes(
            ["query", "reference_output"],
            [["q1", "expected answer"]],
        )
        cases = parse_evaluation_cases_from_excel("test.xlsx", raw)
        assert cases[0]["label"]["answer"] == "expected answer"

    def test_answer_is_optional(self):
        """answer column is optional — empty cell is fine."""
        raw = _make_xlsx_bytes(
            ["query", "answer"],
            [["q1", ""]],
        )
        cases = parse_evaluation_cases_from_excel("test.xlsx", raw)
        assert len(cases) == 1
        assert cases[0]["label"] == {}

    def test_multi_turn_sessions(self):
        """Two cases in same session with different turn orders."""
        raw = _make_xlsx_bytes(
            ["session_id", "request_id", "query", "answer"],
            [
                ["s1", "1", "first question", "first answer"],
                ["s1", "2", "follow-up question", "follow-up answer"],
                ["s2", "1", "new session question", "new session answer"],
            ],
        )
        cases = parse_evaluation_cases_from_excel("test.xlsx", raw)
        assert len(cases) == 3
        assert cases[0]["inputs"]["session_id"] == "s1"
        assert cases[0]["inputs"]["request_id"] == "1"
        assert cases[1]["inputs"]["session_id"] == "s1"
        assert cases[1]["inputs"]["request_id"] == "2"
        assert cases[2]["inputs"]["session_id"] == "s2"

    def test_dropped_columns_are_now_unrecognized(self):
        """custom_variables / case_id / 序号 are no longer recognized."""
        raw = _make_xlsx_bytes(
            ["case_id", "序号", "custom_variables", "query"],
            [["c1", "x1", "{}", "q1"]],
        )
        # query is present so the file parses; the other three columns
        # are silently ignored.
        cases = parse_evaluation_cases_from_excel("test.xlsx", raw)
        assert len(cases) == 1
        assert cases[0]["inputs"]["query"] == "q1"
        assert "case_id" not in cases[0]
        assert "custom_variables" not in cases[0]["inputs"]

    def test_template_round_trip_zh(self):
        """Generate a zh template, fill data, and parse it back."""
        template_bytes = build_evaluation_set_excel_template_bytes(language="zh")
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(template_bytes))
        ws = wb.active
        # Data starts at row 3; clear rows 3+ and add our own
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                cell.value = None
        ws.cell(row=3, column=1, value="rt-session")
        ws.cell(row=3, column=2, value="1")
        ws.cell(row=3, column=3, value="test query")
        ws.cell(row=3, column=4, value="test answer")

        buf = io.BytesIO()
        wb.save(buf)
        raw = buf.getvalue()

        cases = parse_evaluation_cases_from_excel("test.xlsx", raw)
        assert len(cases) >= 1
        assert cases[0]["inputs"]["query"] == "test query"
        assert cases[0]["inputs"]["session_id"] == "rt-session"
        assert cases[0]["label"]["answer"] == "test answer"

    def test_template_round_trip_en(self):
        """Generate an en template, fill data, and parse it back."""
        template_bytes = build_evaluation_set_excel_template_bytes(language="en")
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(template_bytes))
        ws = wb.active
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                cell.value = None
        ws.cell(row=3, column=1, value="en-session")
        ws.cell(row=3, column=2, value="1")
        ws.cell(row=3, column=3, value="en query")
        ws.cell(row=3, column=4, value="en answer")

        buf = io.BytesIO()
        wb.save(buf)
        raw = buf.getvalue()

        cases = parse_evaluation_cases_from_excel("test.xlsx", raw)
        assert len(cases) >= 1
        assert cases[0]["inputs"]["query"] == "en query"
        assert cases[0]["inputs"]["session_id"] == "en-session"


# ---------------------------------------------------------------------------
# parse_evaluation_cases_from_excel — .xlsx path
# ---------------------------------------------------------------------------

class TestParseXlsx:
    def test_empty_file_raises(self):
        from openpyxl import Workbook

        wb = Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError, match="no header row"):
            parse_evaluation_cases_from_excel("test.xlsx", buf.getvalue())

    def test_no_cases_raises(self):
        raw = _make_xlsx_bytes(["query", "answer"], [])
        with pytest.raises(ValueError, match="no cases"):
            parse_evaluation_cases_from_excel("test.xlsx", raw)

    def test_row_missing_query_raises(self):
        raw = _make_xlsx_bytes(
            ["query", "answer"],
            [["has query", "has answer"], ["", "has answer"]],
        )
        with pytest.raises(ValueError, match="query is required"):
            parse_evaluation_cases_from_excel("test.xlsx", raw)

    def test_row_missing_answer_is_ok(self):
        """answer is optional; row with only query should parse fine."""
        raw = _make_xlsx_bytes(
            ["query", "answer"],
            [["has query", ""]],
        )
        cases = parse_evaluation_cases_from_excel("test.xlsx", raw)
        assert len(cases) == 1
        assert cases[0]["inputs"]["query"] == "has query"
        assert cases[0]["label"] == {}

    def test_empty_row_is_skipped(self):
        raw = _make_xlsx_bytes(
            ["query", "answer"],
            [["q1", "a1"], [None, None], ["q2", "a2"]],
        )
        cases = parse_evaluation_cases_from_excel("test.xlsx", raw)
        assert len(cases) == 2

    def test_multiple_rows_parsed_correctly(self):
        raw = _make_xlsx_bytes(
            ["query", "answer"],
            [
                ["first question", "first answer"],
                ["second question", "second answer"],
            ],
        )
        cases = parse_evaluation_cases_from_excel("test.xlsx", raw)
        assert len(cases) == 2
        assert cases[0]["inputs"]["query"] == "first question"
        assert cases[1]["inputs"]["query"] == "second question"
        assert cases[0]["label"]["answer"] == "first answer"

    def test_filename_case_insensitive(self):
        raw = _make_xlsx_bytes(["query", "answer"], [["q1", "a1"]])
        cases = parse_evaluation_cases_from_excel("test.XLSX", raw)
        assert len(cases) == 1


# ---------------------------------------------------------------------------
# parse_evaluation_cases_from_excel — .xls (legacy xlrd) path
# ---------------------------------------------------------------------------

class TestParseXls:
    def test_no_header_row_raises(self):
        with pytest.raises(Exception):
            parse_evaluation_cases_from_excel("test.xls", b"")

    def test_row_missing_query_raises(self):
        raw = _make_xlsx_bytes(
            ["query", "answer"],
            [
                ["has query", "has answer"],
                ["", "has answer"],
            ],
        )
        with pytest.raises(ValueError, match="query is required"):
            parse_evaluation_cases_from_excel("test.xls", raw)

    def test_row_missing_answer_is_ok(self):
        """answer is optional."""
        raw = _make_xlsx_bytes(
            ["query", "answer"],
            [["has query", ""]],
        )
        cases = parse_evaluation_cases_from_excel("test.xls", raw)
        assert len(cases) == 1
        assert cases[0]["inputs"]["query"] == "has query"
        assert cases[0]["label"] == {}

    def test_no_cases_raises(self):
        raw = _make_xlsx_bytes(["query", "answer"], [])
        with pytest.raises(ValueError, match="no cases"):
            parse_evaluation_cases_from_excel("test.xls", raw)

    def test_multiple_rows(self):
        raw = _make_xlsx_bytes(
            ["query", "answer"],
            [["q1", "a1"], ["q2", "a2"]],
        )
        cases = parse_evaluation_cases_from_excel("test.xls", raw)
        assert len(cases) == 2
        assert [c["inputs"]["query"] for c in cases] == ["q1", "q2"]

    def test_filename_case_insensitive(self):
        raw = _make_xlsx_bytes(["query", "answer"], [["q1", "a1"]])
        cases = parse_evaluation_cases_from_excel("test.XLS", raw)
        assert len(cases) == 1

    def test_chinese_headers_xls(self):
        """xls path should also handle Chinese headers."""
        raw = _make_xlsx_bytes(
            ["会话ID", "请求顺序", "问题", "答案"],
            [["s1", "1", "你好", "世界"]],
        )
        cases = parse_evaluation_cases_from_excel("test.xls", raw)
        assert len(cases) == 1
        assert cases[0]["inputs"]["query"] == "你好"
        assert cases[0]["inputs"]["session_id"] == "s1"
        assert cases[0]["label"]["answer"] == "世界"
