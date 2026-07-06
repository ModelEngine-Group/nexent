import asyncio
import io
import json
import logging
from statistics import mean
from typing import Any, Dict, List, Optional, Tuple

from adapters.exception import JiuwenSDKError, JiuwenSDKUnavailableError

try:
    from adapters.jiuwen_sdk_adapter import JiuwenSDKAdapter
except ModuleNotFoundError:
    JiuwenSDKAdapter = None  # type: ignore[assignment, misc]
from consts.model import AgentRequest
from database.agent_evaluation_db import (
    create_agent_evaluation,
    create_agent_evaluation_cases,
    get_agent_evaluation,
    list_agent_evaluation_cases,
    list_agent_evaluations_by_agent,
    soft_delete_agent_evaluation,
    update_agent_evaluation_case_result,
    update_agent_evaluation_status,
)
from database.evaluation_set_db import get_evaluation_set_cases_all
from services.evaluation_set_service import resolve_latest_published_version_no
from services.agent_service import prepare_agent_run
from utils.thread_utils import pool
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re

logger = logging.getLogger(__name__)


# Log records emitted during SDK invocations may bleed into the ``reason``
# field as ``"[HH:MM:SS LEVEL logger_name] {...payload...}"``. Extract the
# embedded JSON ``reason`` from those polluted strings so the report shows the
# judge's actual explanation rather than the surrounding log envelope.
_LOG_PREFIX_RE = re.compile(
    r"(?:"
    # Short form: ``[HH:MM:SS LEVEL logger] ``
    r"\[(\d{2}:\d{2}:\d{2}|\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\s+\w+\s+[\w.]+\]"
    # Long form: ``YYYY-MM-DD HH:MM:SS | logger_name | trace_id | LEVEL | ``
    r"|(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})"
    r"(?:\s*\|\s*[\w.]+){1,3}\s*\|?"
    r")\s*"
)
_MARKDOWN_FENCE_RE = re.compile(r"```(?:json)?\s*(\{.*?\})\s*```", re.DOTALL)


def _extract_clean_reason(raw: Any) -> str:
    """Best-effort extraction of the judge model's reason text.

    The ``reason`` column persisted by the judge pipeline can take three
    shapes depending on which step produced the value:

    1. The plain judge verdict: ``"pass"`` or ``"fail"``.
    2. The judge's free-form explanation (when the SDK upgrades to return it).
    3. A captured log record of the form
       ``"[HH:MM:SS LEVEL llm] {event_id:..., response_content: '```json\\n{result, reason}\\n```', ...}"``
       where the judge JSON is nested inside ``response_content``.

    This helper unwraps case 3 by parsing the outer JSON, pulling
    ``response_content``, stripping the markdown code fence, and returning the
    inner ``reason`` field. Cases 1 and 2 are returned as-is.
    """
    if raw is None:
        return ""
    text = str(raw).strip()
    if not text:
        return ""

    # Strip leading "[HH:MM:SS LEVEL logger] " log envelope(s) that may have
    # been prepended one or more times by the root-logger StreamHandler when
    # the SDK redirected ``response`` into a log call.
    stripped = _LOG_PREFIX_RE.sub("", text).strip()
    if not stripped:
        return text

    # Try the full string first; if it is not parseable JSON, fall back to
    # the inner match. This handles both well-formed outer objects and strings
    # that have stray characters around a single JSON object.
    parsed: Optional[Dict[str, Any]] = None
    try:
        parsed = json.loads(stripped)
    except (ValueError, TypeError):
        match = _JSON_OBJECT_RE.search(stripped)
        if match:
            try:
                parsed = json.loads(match.group(0))
            except (ValueError, TypeError):
                parsed = None

    if not isinstance(parsed, dict):
        # Not a JSON envelope at all — leave the stripped text in place so
        # plain "pass"/"fail" and free-form explanations still render.
        return stripped

    response_content = parsed.get("response_content")
    if isinstance(response_content, str):
        # Strip the ```json ... ``` fence the judge model wraps its verdict in.
        fence_match = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", response_content, re.DOTALL)
        if fence_match:
            try:
                inner = json.loads(fence_match.group(1))
                if isinstance(inner, dict) and isinstance(inner.get("reason"), str):
                    return inner["reason"].strip()
            except (ValueError, TypeError):
                pass

    reason_field = parsed.get("reason")
    if isinstance(reason_field, str):
        return reason_field.strip()

    return stripped


def _iter_log_envelopes(text: str):
    """Yield ``(log_prefix, json_payload)`` for every
    ``"[HH:MM:SS LEVEL logger] {...}"`` envelope in *text*.

    The SDK's root logger may emit several records per evaluation
    (``"OpenAI API response received"``, ``"Before parse response with output
    parser"``, ``"Before parse content with parser"``, ...). Some of them carry
    the judge's verdict inside their JSON payload, others do not. The caller
    picks the envelope that contains ``result`` + ``reason``.
    """
    cursor = 0
    while cursor < len(text):
        # Skip any whitespace / newlines separating envelopes.
        while cursor < len(text) and text[cursor] in " \t\r\n":
            cursor += 1
        if cursor >= len(text):
            break
        match = _LOG_PREFIX_RE.match(text, cursor)
        if not match:
            break
        prefix_end = match.end()
        # Find the balanced JSON object that begins at prefix_end. If the
        # payload is not valid JSON (e.g. truncated) we still want to advance
        # the cursor past the broken prefix so the loop can find the next
        # envelope.
        depth = 0
        payload_start = -1
        payload_end = -1
        in_string = False
        escape = False
        for idx in range(prefix_end, len(text)):
            ch = text[idx]
            if in_string:
                if escape:
                    escape = False
                elif ch == "\\":
                    escape = True
                elif ch == '"':
                    in_string = False
                continue
            if ch == '"':
                in_string = True
                continue
            if ch == "{":
                if depth == 0:
                    payload_start = idx
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0 and payload_start != -1:
                    payload_end = idx + 1
                    break
        if payload_start == -1 or payload_end == -1:
            # No balanced JSON after this prefix — skip the prefix and keep
            # scanning for the next log envelope.
            cursor = prefix_end
            continue
        yield text[cursor:prefix_end], text[payload_start:payload_end]
        cursor = payload_end


def _reason_from_json_envelope(payload: str) -> Optional[str]:
    """Pull the judge's ``reason`` text out of one log-envelope JSON payload.

    Tries the following strategies, in order:

    1. ``payload["response_content"]`` wrapped in a markdown code fence — the
       shape produced by the judge LLM's first ``llm_call_end`` event.
    2. ``payload["response_content"]`` being a plain JSON string.
    3. ``payload["response"]["choices"][0]["message"]["content"]`` — the
       OpenAI ``ChatCompletion`` repr captured by the request-side log.
    4. ``payload["reason"]`` at the top level — fallback.
    """
    try:
        data = json.loads(payload)
    except (ValueError, TypeError):
        return None
    if not isinstance(data, dict):
        return None

    response_content = data.get("response_content")
    if isinstance(response_content, str):
        fence_match = _MARKDOWN_FENCE_RE.search(response_content)
        if fence_match:
            try:
                inner = json.loads(fence_match.group(1))
            except (ValueError, TypeError):
                inner = None
            if isinstance(inner, dict):
                reason = inner.get("reason")
                if isinstance(reason, str):
                    return reason.strip()
        # ``response_content`` may be the raw JSON string itself.
        try:
            inner = json.loads(response_content)
        except (ValueError, TypeError):
            inner = None
        if isinstance(inner, dict):
            reason = inner.get("reason")
            if isinstance(reason, str):
                return reason.strip()

    # OpenAI ChatCompletion repr — unwrap the ``message.content`` field.
    # The SDK captures the OpenAI response by calling ``repr(chat_completion)``
    # so the ``response`` field stores a Python repr (single-quoted strings,
    # ``\n`` literal characters, ...) rather than a JSON document. Use a
    # targeted regex to pull the ``content='...'`` blob out of the repr, then
    # unwrap the markdown fence the judge model wraps its verdict in.
    # The field can live either at the top level (``payload["response"]``) or
    # nested under ``payload["metadata"]["response"]`` depending on which SDK
    # code path emitted the log record.
    response = data.get("response")
    if response is None:
        metadata = data.get("metadata")
        if isinstance(metadata, dict):
            response = metadata.get("response")
    if isinstance(response, str) and "ChatCompletion" in response:
        # Use a non-greedy match that ends at the next ``, refusal`` token.
        # The repr is single-quoted and the ``content`` field is followed by
        # ``, refusal=`` so this anchor avoids trying to balance quotes in
        # the repr string — which is fragile once ``\n`` has been unescaped
        # by ``json.loads``.
        content_match = re.search(
            r"ChatCompletionMessage\(content='(.*?)', refusal=",
            response,
            re.DOTALL,
        )
        if content_match:
            # The captured group is the raw repr body; turn the literal
            # ``\n`` / ``\"`` / ``\\`` escape sequences back into real
            # characters, then look for the markdown code fence inside.
            try:
                content = content_match.group(1).encode("utf-8").decode("unicode_escape")
            except UnicodeDecodeError:
                content = content_match.group(1)
            if isinstance(content, str):
                fence_match = _MARKDOWN_FENCE_RE.search(content)
                if fence_match:
                    try:
                        inner = json.loads(fence_match.group(1))
                    except (ValueError, TypeError):
                        inner = None
                    if isinstance(inner, dict):
                        reason = inner.get("reason")
                        if isinstance(reason, str):
                            return reason.strip()
                # The judge verdict may live directly on ``response`` if the
                # repr was stored without the ``metadata.response`` wrapping.
                try:
                    inner = json.loads(content)
                except (ValueError, TypeError):
                    inner = None
                if isinstance(inner, dict):
                    reason = inner.get("reason")
                    if isinstance(reason, str):
                        return reason.strip()
    elif isinstance(response, str):
        # Some SDKs serialise the response as a plain JSON string.
        try:
            response_obj = json.loads(response)
        except (ValueError, TypeError):
            response_obj = None
        if isinstance(response_obj, dict):
            choices = response_obj.get("choices")
            if isinstance(choices, list) and choices:
                first = choices[0]
                if isinstance(first, dict):
                    message = first.get("message")
                    if isinstance(message, dict):
                        content = message.get("content")
                        if isinstance(content, str):
                            fence_match = _MARKDOWN_FENCE_RE.search(content)
                            if fence_match:
                                try:
                                    inner = json.loads(fence_match.group(1))
                                except (ValueError, TypeError):
                                    inner = None
                                if isinstance(inner, dict):
                                    reason = inner.get("reason")
                                    if isinstance(reason, str):
                                        return reason.strip()

    top_reason = data.get("reason")
    if isinstance(top_reason, str):
        return top_reason.strip()

    return None


def _extract_clean_reason_v2(raw: Any) -> str:
    """Walk every log envelope in *raw* and pull out the judge reason.

    See ``_extract_clean_reason`` for the original entry point; this helper
    handles the noisy multi-envelope shape where several SDK log records
    (with the verdict buried in the first one) are concatenated.
    """
    if raw is None:
        return ""
    text = str(raw).strip()
    if not text:
        return ""

    # Fast path: the string is already a single JSON object (no log envelope)
    # containing the judge verdict.
    standalone = _reason_from_json_envelope(text)
    if standalone is not None:
        return standalone

    # Slow path: walk every ``[time LEVEL logger] {...}`` envelope and take
    # the first one whose payload yields a reason.
    for _, payload in _iter_log_envelopes(text):
        reason = _reason_from_json_envelope(payload)
        if reason is not None:
            return reason

    # No JSON envelope produced a reason. Return the input with leading log
    # prefixes stripped so plain "pass" / "fail" and free-form explanations
    # render without the surrounding noise.
    stripped = text
    while True:
        match = _LOG_PREFIX_RE.match(stripped)
        if not match:
            break
        stripped = stripped[match.end():].lstrip()
    return stripped or text


def _is_llm_related_error(exc: Exception) -> bool:
    """Check if an exception is related to LLM API calls."""
    error_str = str(exc).lower()
    llm_keywords = [
        "openai", "api", "llm", "model", "completion", "chat",
        "connection", "timeout", "rate limit", "authentication",
        "invalid response", "async invoke", "jiuwen", "sdk",
        "schedule new futures", "interpreter shutdown",
    ]
    return any(keyword in error_str for keyword in llm_keywords)


def _generate_friendly_error_message(exc: Exception, default_msg: str) -> str:
    """Generate a friendly error message for LLM-related errors using another LLM."""
    if not _is_llm_related_error(exc):
        return default_msg

    # Only call LLM for LLM-related errors
    try:
        import os
        from openai import AsyncOpenAI

        client = AsyncOpenAI(api_key=os.environ.get("OPENAI_API_KEY"))
        error_snippet = str(exc)[:500]

        response = asyncio.run(
            client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "You are a helpful assistant that explains technical errors "
                            "to end users in simple Chinese. Be concise and actionable. "
                            "Focus on what the user can do to fix the problem."
                        ),
                    },
                    {
                        "role": "user",
                        "content": (
                            f"以下是智能体评估时的错误信息，请用简单的中文解释给用户，并给出建议的操作：\n\n{error_snippet}"
                        ),
                    },
                ],
                max_tokens=200,
                temperature=0.3,
            )
        )
        return response.choices[0].message.content or default_msg
    except Exception as llm_exc:
        logger.warning("Failed to generate friendly error message: %r", llm_exc)
        return default_msg

logger = logging.getLogger("agent_evaluation_service")


def _make_background_done_callback(
    tenant_id: str,
    user_id: str,
    agent_evaluation_id: int,
):
    """Return a done-callback that marks the run FAILED if the worker raised."""

    def callback(future):
        exc = future.exception()
        if exc is not None:
            logger.exception(
                "Background evaluation run failed (id=%s): %r",
                agent_evaluation_id,
                exc,
            )
            friendly_msg = _generate_friendly_error_message(exc, str(exc))
            try:
                update_agent_evaluation_status(
                    agent_evaluation_id=agent_evaluation_id,
                    tenant_id=tenant_id,
                    status="FAILED",
                    updated_by=user_id,
                    error_message=friendly_msg,
                )
            except Exception as update_exc:
                logger.error(
                    "Failed to write FAILED status for evaluation id=%s: %r",
                    agent_evaluation_id,
                    update_exc,
                )

    return callback


def _build_case_for_jiuwen(
    inputs: Dict[str, Any],
    label: Dict[str, Any],
) -> Dict[str, Any]:
    # Jiuwen Case(inputs, label) will accept arbitrary keys, but we keep the stable schema
    return {
        "inputs": {
            "query": inputs.get("query", ""),
        },
        "label": {"answer": label.get("answer", "")},
    }


async def _run_agent_to_final_answer(
    agent_id: int,
    tenant_id: str,
    user_id: str,
    query: str,
    version_no: int,
) -> str:
    """Run agent once and aggregate final answer text."""

    # Build a single-turn AgentRequest. We do not persist messages for offline eval.
    agent_request = AgentRequest(
        query=query,
        conversation_id=0,
        history=None,
        minio_files=None,
        agent_id=agent_id,
        version_no=version_no,
        is_debug=True,
    )

    agent_run_info, memory_context = await prepare_agent_run(
        agent_request=agent_request,
        user_id=user_id,
        tenant_id=tenant_id,
        allow_memory_search=False,
    )

    # Stream chunks from core agent runner and extract final_answer content.
    from nexent.core.agents.run_agent import agent_run

    final_answer_parts: List[str] = []
    async for chunk in agent_run(agent_run_info):
        try:
            if isinstance(chunk, str):
                data = json.loads(chunk)
                if isinstance(data, dict) and data.get("type") == "final_answer":
                    content = data.get("content")
                    if isinstance(content, str):
                        final_answer_parts.append(content)
        except Exception:
            continue

    return "".join(final_answer_parts).strip()


def create_agent_evaluation_run_impl(
    tenant_id: str,
    user_id: str,
    agent_id: int,
    evaluation_set_id: int,
    judge_model_id: int,
) -> Dict[str, Any]:
    set_cases = get_evaluation_set_cases_all(evaluation_set_id=evaluation_set_id, tenant_id=tenant_id)
    if not set_cases:
        raise ValueError("evaluation set has no cases")

    agent_version_no = resolve_latest_published_version_no(agent_id=agent_id, tenant_id=tenant_id)

    run = create_agent_evaluation(
        tenant_id=tenant_id,
        agent_id=agent_id,
        agent_version_no=agent_version_no,
        evaluation_set_id=evaluation_set_id,
        total=len(set_cases),
        judge_model_id=judge_model_id,
        created_by=user_id,
    )

    create_agent_evaluation_cases(
        tenant_id=tenant_id,
        agent_evaluation_id=run["agent_evaluation_id"],
        set_cases=set_cases,
        created_by=user_id,
    )

    # Kick off background execution — attach a callback so that any uncaught
    # exception inside the worker thread is surfaced as a RUN→FAILED transition
    # instead of silently leaving the run stuck at RUNNING forever.
    future = pool.submit(
        execute_agent_evaluation_run,
        tenant_id,
        user_id,
        run["agent_evaluation_id"],
        judge_model_id,
    )
    future.add_done_callback(
        _make_background_done_callback(
            tenant_id,
            user_id,
            run["agent_evaluation_id"],
        )
    )

    return run


def execute_agent_evaluation_run(
    tenant_id: str,
    user_id: str,
    agent_evaluation_id: int,
    judge_model_id: Optional[int] = None,
):
    """Background execution entry point (sync).

    ``judge_model_id`` is normally supplied by ``submit`` when the run is
    first created. If the worker process restarts mid-run and the queued
    payload is lost, we fall back to whatever was persisted on the run
    record so the run can still recover.
    """
    try:
        update_agent_evaluation_status(
            agent_evaluation_id=agent_evaluation_id,
            tenant_id=tenant_id,
            status="RUNNING",
            updated_by=user_id,
        )

        run = get_agent_evaluation(agent_evaluation_id=agent_evaluation_id, tenant_id=tenant_id)
        agent_id = int(run["agent_id"])
        agent_version_no = int(run["agent_version_no"])
        if judge_model_id is None:
            judge_model_id = run.get("judge_model_id")
        if judge_model_id is None:
            raise ValueError("judge_model_id is required but neither passed in nor persisted on the run")
        judge_model_id = int(judge_model_id)

        if JiuwenSDKAdapter is None:
            raise JiuwenSDKUnavailableError("Jiuwen SDK adapter is unavailable. Please install optional dependencies for openjiuwen.")

        adapter = JiuwenSDKAdapter(model_id=judge_model_id, tenant_id=tenant_id)

        cases = list_agent_evaluation_cases(agent_evaluation_id=agent_evaluation_id, tenant_id=tenant_id, limit=100000, offset=0)
        scores: List[float] = []

        for idx, c in enumerate(cases, start=1):
            case_id = c["agent_evaluation_case_id"]
            update_agent_evaluation_case_result(
                agent_evaluation_case_id=case_id,
                tenant_id=tenant_id,
                status="RUNNING",
                updated_by=user_id,
            )

            inputs = c["inputs"] or {}
            label = c["label"] or {}

            query = inputs.get("query", "")

            try:
                answer_text = asyncio.run(
                    _run_agent_to_final_answer(
                        agent_id=agent_id,
                        tenant_id=tenant_id,
                        user_id=user_id,
                        query=query,
                        version_no=agent_version_no,
                    )
                )

                predict = {"answer": answer_text}

                # Judge with openjiuwen LLM-as-judge metric (binary 1/0)
                expected = (label.get("answer") or "").strip()

                score, reason = adapter.evaluate_semantic_consistency(
                    question=query,
                    expected_answer=expected,
                    model_answer=answer_text,
                )

                update_agent_evaluation_case_result(
                    agent_evaluation_case_id=case_id,
                    tenant_id=tenant_id,
                    status="COMPLETED",
                    predict=predict,
                    score=score,
                    pass_status="pass" if score == 1 else "fail",
                    reason=reason,
                    updated_by=user_id,
                )

                scores.append(score)

            except Exception as exc:
                logger.exception("Evaluation case failed: %r", exc)
                friendly_msg = _generate_friendly_error_message(exc, str(exc))
                update_agent_evaluation_case_result(
                    agent_evaluation_case_id=case_id,
                    tenant_id=tenant_id,
                    status="FAILED",
                    pass_status="fail",
                    error_message=friendly_msg,
                    updated_by=user_id,
                )

            update_agent_evaluation_status(
                agent_evaluation_id=agent_evaluation_id,
                tenant_id=tenant_id,
                status="RUNNING",
                updated_by=user_id,
                progress_done=idx,
            )

        overall = float(mean(scores)) if scores else 0.0
        update_agent_evaluation_status(
            agent_evaluation_id=agent_evaluation_id,
            tenant_id=tenant_id,
            status="COMPLETED",
            updated_by=user_id,
            score_overall=overall,
            progress_done=len(cases),
        )

    except Exception as exc:
        logger.exception("Evaluation run failed: %r", exc)
        friendly_msg = _generate_friendly_error_message(exc, str(exc))
        update_agent_evaluation_status(
            agent_evaluation_id=agent_evaluation_id,
            tenant_id=tenant_id,
            status="FAILED",
            updated_by=user_id,
            error_message=friendly_msg,
        )


def get_agent_evaluation_run_impl(agent_evaluation_id: int, tenant_id: str) -> Dict[str, Any]:
    return get_agent_evaluation(agent_evaluation_id=agent_evaluation_id, tenant_id=tenant_id)


def list_agent_evaluations_by_agent_impl(
    agent_id: int,
    tenant_id: str,
    limit: int = 50,
    offset: int = 0,
) -> List[Dict[str, Any]]:
    return list_agent_evaluations_by_agent(agent_id=agent_id, tenant_id=tenant_id, limit=limit, offset=offset)


def list_agent_evaluation_cases_impl(
    agent_evaluation_id: int,
    tenant_id: str,
    limit: int = 50,
    offset: int = 0,
) -> List[Dict[str, Any]]:
    return list_agent_evaluation_cases(agent_evaluation_id=agent_evaluation_id, tenant_id=tenant_id, limit=limit, offset=offset)


def delete_agent_evaluation_run_impl(
    agent_evaluation_id: int,
    tenant_id: str,
    user_id: str,
) -> None:
    """Soft-delete an evaluation run.

    Only the creator can delete the run. Tenant admins are handled at the
    service layer when permissions are extended.
    """
    run = get_agent_evaluation(agent_evaluation_id=agent_evaluation_id, tenant_id=tenant_id)
    if run.get("created_by") != user_id:
        raise ValueError("Only the creator can delete this evaluation run")
    soft_delete_agent_evaluation(agent_evaluation_id, tenant_id, user_id)


def generate_agent_evaluation_report_impl(
    agent_evaluation_id: int,
    tenant_id: str,
) -> Tuple[bytes, int]:
    """Build the evaluation report workbook.

    Returns ``(excel_bytes, failed_count)``. ``failed_count`` lets the API
    layer name the downloaded file ``_all.xlsx`` vs ``_failed.xlsx`` so a
    clean run does not download a file whose name suggests failure.
    """
    run = get_agent_evaluation(agent_evaluation_id=agent_evaluation_id, tenant_id=tenant_id)
    all_cases = list_agent_evaluation_cases(
        agent_evaluation_id=agent_evaluation_id,
        tenant_id=tenant_id,
        limit=100000,
        offset=0,
    )

    failed_cases = [
        c for c in all_cases
        if c.get("status") == "FAILED"
        or c.get("score") == 0
        or c.get("pass_status") == "fail"
    ]
    pass_count = sum(
        1 for c in all_cases
        if c.get("status") != "FAILED" and c.get("score") == 1
    )
    fail_count = len(failed_cases)
    total = len(all_cases)
    pass_rate = f"{(pass_count / total * 100):.2f}%" if total else "-"

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "概要"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Excel's default cell style is right-aligned for numbers; force every
    # value cell to render flush-left so mixed types (numbers, percentages,
    # timestamps, error messages) line up consistently in the report.
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # The three "ID" rows (评估ID / Agent ID / Evaluation Set ID) expose the
    # human-readable name (judge model / agent / evaluation set) instead of
    # the opaque numeric ID. Falls back to the raw ID when the name lookup
    # returned nothing, so the cell is never blank.
    judge_model_label = run.get("judge_model_name") or run.get("judge_model_id") or "-"
    agent_label = run.get("agent_name") or run.get("agent_id") or "-"
    evaluation_set_label = run.get("evaluation_set_name") or run.get("evaluation_set_id") or "-"

    summary_rows = [
        ("测评模型", judge_model_label),
        ("智能体名称", agent_label),
        ("智能体版本", run.get("agent_version_no", "")),
        ("评测集名称", evaluation_set_label),
        ("状态", run.get("status", "")),
        ("用例总数", total),
        ("通过用例数", pass_count),
        ("失败用例数", fail_count),
        ("通过率", pass_rate),
        ("综合得分", f"{run.get('score_overall', 0):.4f}" if run.get('score_overall') is not None else "-"),
        ("错误信息", run.get("error_message") or "-"),
        ("创建时间", run.get("create_time") or "-"),
        ("报告范围", "失败用例"),
    ]

    ws_summary.append(["字段", "值"])
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Pre-build the data rows so we can apply the left alignment to every
    # value cell in one pass (rows start at index 3 after the header in row 1
    # and the empty default row that ``append`` does not create; openpyxl
    # places our header at row 1 and the data rows begin at row 2).
    for field, value in summary_rows:
        ws_summary.append([field, value])

    for row_idx in range(2, 2 + len(summary_rows)):
        ws_summary.cell(row=row_idx, column=1).alignment = left_align
        ws_summary.cell(row=row_idx, column=2).alignment = left_align

    ws_summary.column_dimensions["A"].width = 24
    ws_summary.column_dimensions["B"].width = 60

    ws_cases = wb.create_sheet("失败用例")

    # All columns in the failed-cases sheet are localized to Chinese to match
    # the rest of the report.
    case_headers = [
        "用例ID",
        "问题",
        "期望答案",
        "模型答案",
        "得分",
        "评测理由",
        "用例状态",
        "错误信息",
    ]
    ws_cases.append(case_headers)
    for cell in ws_cases[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for c in failed_cases:
        inputs = c.get("inputs") or {}
        label = c.get("label") or {}
        predict = c.get("predict") or {}
        score = c.get("score")
        if score == 0:
            score_str = "0.0000"
        elif isinstance(score, (int, float)):
            score_str = f"{score:.4f}"
        elif score is not None:
            score_str = str(score)
        else:
            score_str = "-"

        # error_message is only meaningful when the case is FAILED
        # (execution crashed before/during model call). For scored cases
        # that simply scored 0 it is usually empty.
        error_msg = c.get("error_message") or ""

        ws_cases.append([
            c.get("agent_evaluation_case_id", ""),
            inputs.get("query", ""),
            label.get("answer", ""),
            predict.get("answer", ""),
            score_str,
            _extract_clean_reason_v2(c.get("reason")),
            c.get("status", ""),
            error_msg,
        ])

    # Apply left alignment + wrap to every data cell so the row stays readable
    # when any column overflows the configured width.
    if failed_cases:
        last_data_row = 1 + len(failed_cases)
        for row in ws_cases.iter_rows(
            min_row=2, max_row=last_data_row, min_col=1, max_col=len(case_headers)
        ):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    widths = {"A": 14, "B": 50, "C": 40, "D": 40, "E": 10, "F": 50, "G": 12, "H": 40}
    for col, w in widths.items():
        ws_cases.column_dimensions[col].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), fail_count
