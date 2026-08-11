"""Excel template generation and case parsing for evaluation sets.

Supports exactly four columns, in either Chinese or English:

  - session_id / 会话ID
  - request_id / 请求顺序 / turn_order
  - query / 问题            (required)
  - answer / 答案 / reference_output

The .xlsx and .xls parsers share a single code path: ``_load_rows`` returns
a list of row tuples plus a uniform cell accessor, so the rest of the
parsing logic is identical for both formats.
"""

import io
from typing import Any

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ── Column layout (4 fields, bilingual headers) ────────────────────

TEMPLATE_HEADERS: dict[str, list[str]] = {
    "zh": ["会话ID", "请求顺序", "问题", "答案"],
    "en": ["session_id", "request_id", "query", "answer"],
}

INSTRUCTIONS: dict[str, list[str]] = {
    "zh": [
        "会话ID（多轮对话填写）\n同一会话的多轮使用相同ID\n不同编号表示不同会话\n单轮对话可留空",
        "请求顺序（多轮对话填写）\n标记每轮对话的顺序\n同一会话内从1递增\n单轮对话可留空",
        "问题（必填*）\n用户输入内容\n不能为空",
        "答案（选填）\n参考答案\n用于后续打分或标注",
    ],
    "en": [
        "Session ID (multi-turn)\nSame ID for all turns in one conversation\nDifferent ID = a new conversation\nCan be empty for single-turn",
        "Request ID (multi-turn)\nMarks the order of each turn\nIncrements from 1 within a session\nCan be empty for single-turn",
        "Query (required*)\nUser input content\nCannot be empty",
        "Answer (optional)\nReference answer\nUsed for scoring or annotation",
    ],
}

EXAMPLE_ROWS: dict[str, list[list[str]]] = {
    "zh": [
        ["s1", "1", "1+1等于几？", "2"],
        ["s1", "2", "再乘以3呢？", "6"],
        ["s2", "1", "中国首都是哪里？", "北京"],
    ],
    "en": [
        ["s1", "1", "What is 1+1?", "2"],
        ["s2", "1", "What is the capital of France?", "Paris"],
        ["s2", "2", "What is its population?", "About 2.1 million"],
    ],
}

# Canonical field names used inside the parser.
FIELD_SESSION_ID = "session_id"
FIELD_REQUEST_ID = "request_id"
FIELD_QUERY = "query"
FIELD_ANSWER = "answer"

REQUIRED_FIELDS: tuple[str, ...] = (FIELD_QUERY,)
ALL_FIELDS: tuple[str, ...] = (
    FIELD_SESSION_ID,
    FIELD_REQUEST_ID,
    FIELD_QUERY,
    FIELD_ANSWER,
)

# Header alias map: any recognized header (English or Chinese, possibly
# with a trailing "*" required marker) maps to one of the canonical names.
# ASCII keys are matched case-insensitively; Chinese keys must match exactly
# after stripping whitespace.
HEADER_ALIASES: dict[str, str] = {
    # session_id
    "session_id": FIELD_SESSION_ID,
    "sessionid": FIELD_SESSION_ID,
    "会话id": FIELD_SESSION_ID,
    "会话ID": FIELD_SESSION_ID,
    # request_id (also accept turn_order as an alias)
    "request_id": FIELD_REQUEST_ID,
    "requestid": FIELD_REQUEST_ID,
    "turn_order": FIELD_REQUEST_ID,
    "turnorder": FIELD_REQUEST_ID,
    "turn": FIELD_REQUEST_ID,
    "请求顺序": FIELD_REQUEST_ID,
    # query
    "query": FIELD_QUERY,
    "问题": FIELD_QUERY,
    # answer (also accept reference_output / expected_output as aliases)
    "answer": FIELD_ANSWER,
    "reference_output": FIELD_ANSWER,
    "referenceoutput": FIELD_ANSWER,
    "expected_output": FIELD_ANSWER,
    "expectedoutput": FIELD_ANSWER,
    "答案": FIELD_ANSWER,
}


def _normalize_header(v: Any) -> str:
    """Strip whitespace and lowercase a header cell value.

    Lowercasing is a no-op for pure Chinese strings but lets ASCII headers
    like ``SESSION_ID`` match the lowercase keys in :data:`HEADER_ALIASES`.
    """
    if v is None:
        return ""
    return str(v).strip().lower()


def _canonical_header(v: Any) -> str | None:
    """Map a header cell to its canonical field name, or None if unknown.

    A trailing ``*`` (required marker) is tolerated. ASCII headers are
    matched case-insensitively; Chinese headers must match exactly after
    stripping.
    """
    key = _normalize_header(v).rstrip("*").strip()
    if not key:
        return None
    if key in HEADER_ALIASES:
        return HEADER_ALIASES[key]
    return HEADER_ALIASES.get(key.lower())


# ── Loading: unify .xlsx and .xls into a single row-list API ────────


def _load_xlsx_rows(raw: bytes) -> list[tuple[Any, ...]]:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


def _load_xls_rows(raw: bytes) -> list[tuple[Any, ...]]:
    book = xlrd.open_workbook(file_contents=raw)
    sheet = book.sheet_by_index(0)
    return [tuple(sheet.row_values(r)) for r in range(sheet.nrows)]


def _load_rows(filename: str, raw: bytes) -> list[tuple[Any, ...]]:
    """Return all rows from a .xlsx or .xls file as a list of tuples."""
    lower = (filename or "").lower()
    if lower.endswith(".xlsx"):
        return _load_xlsx_rows(raw)
    if lower.endswith(".xls"):
        return _load_xls_rows(raw)
    raise ValueError("Unsupported file type. Please upload .xlsx or .xls")


def _cell_str(row: tuple[Any, ...], idx: int) -> str | None:
    """Return the stripped string value of a cell, or None if empty/missing.

    ``xlrd`` returns floats for numeric cells; we coerce integer floats to
    ``int`` so ``1`` does not round-trip as ``"1.0"``.
    """
    if idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        s = str(int(v))
    else:
        s = str(v).strip()
    return s if s else None


def _cell_for(
    row: tuple[Any, ...], header_map: dict[str, int], field: str
) -> str | None:
    """Return the stripped cell value for ``field`` in ``row``."""
    idx = header_map.get(field)
    if idx is None:
        return None
    return _cell_str(row, idx)


def _find_header_row(rows: list[tuple[Any, ...]]) -> int:
    """Return the index of the first row that has any recognized column."""
    for r_idx, row in enumerate(rows):
        for v in row:
            if _canonical_header(v) is not None:
                return r_idx
    raise ValueError("Excel contains no header row")


def _build_header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    """Build canonical field name → column index from the header row."""
    header_map: dict[str, int] = {}
    for idx, v in enumerate(header_row):
        canon = _canonical_header(v)
        if canon and canon not in header_map:
            header_map[canon] = idx
    return header_map


def _build_normalized_case(
    row: tuple[Any, ...], header_map: dict[str, int], row_no: int
) -> dict[str, Any] | None:
    """Convert a data row into a normalized case dict.

    Returns ``None`` for fully empty rows. Raises ``ValueError`` if the
    required ``query`` field is missing.
    """
    query = _cell_for(row, header_map, FIELD_QUERY)
    answer = _cell_for(row, header_map, FIELD_ANSWER)
    session_id = _cell_for(row, header_map, FIELD_SESSION_ID)
    request_id = _cell_for(row, header_map, FIELD_REQUEST_ID)

    if not any([query, answer, session_id, request_id]):
        return None

    if not query:
        raise ValueError(f"Row {row_no}: query is required")

    inputs: dict[str, Any] = {"query": query}
    if session_id:
        inputs[FIELD_SESSION_ID] = session_id
    if request_id:
        inputs[FIELD_REQUEST_ID] = request_id

    normalized: dict[str, Any] = {
        "inputs": inputs,
        "label": {"answer": answer} if answer else {},
    }
    if session_id:
        normalized[FIELD_SESSION_ID] = session_id
    if request_id:
        normalized["turn_order"] = request_id
    return normalized


# ── Public API ─────────────────────────────────────────────────────


def parse_evaluation_cases_from_excel(
    filename: str, raw: bytes
) -> list[dict[str, Any]]:
    """Parse evaluation cases from .xlsx or .xls.

    Supported headers (case-insensitive for ASCII; Chinese exact match):

      - ``session_id`` / ``会话ID``
      - ``request_id`` / ``请求顺序`` / ``turn_order``
      - ``query`` / ``问题``           (required)
      - ``answer`` / ``答案`` / ``reference_output``

    Returns normalized case dicts compatible with
    :func:`insert_evaluation_set_cases`.
    """
    rows = _load_rows(filename, raw)
    if not rows:
        raise ValueError("Excel contains no header row")

    header_idx = _find_header_row(rows)
    header_map = _build_header_map(rows[header_idx])

    for field in REQUIRED_FIELDS:
        if field not in header_map:
            raise ValueError(f"Missing required column: {field}")

    cases: list[dict[str, Any]] = []
    for r_idx in range(header_idx + 1, len(rows)):
        case = _build_normalized_case(rows[r_idx], header_map, r_idx + 1)
        if case is None:
            continue
        case["order_no"] = len(cases)
        cases.append(case)

    if not cases:
        raise ValueError("Excel contains no cases")
    return cases


def _apply_template_styles(ws, language: str) -> None:
    """Apply instruction/header styling and column widths to a worksheet.

    ``language`` selects which header set is in row 2 so the required-column
    highlight (the ``query`` / ``问题`` column) is applied to the right cell.
    """
    bold = Font(bold=True)
    instruction_font = Font(italic=True, color="808080")
    instruction_align = Alignment(wrap_text=True, vertical="top")
    required_fill = PatternFill(
        start_color="FFF7E6", end_color="FFF7E6", fill_type="solid"
    )

    headers = TEMPLATE_HEADERS.get(language, TEMPLATE_HEADERS["zh"])
    instructions = INSTRUCTIONS.get(language, INSTRUCTIONS["zh"])
    required_title = "问题" if language == "zh" else "query"

    for col_idx in range(1, len(instructions) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = instruction_font
        cell.alignment = instruction_align
    ws.row_dimensions[1].height = 64

    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = bold
        if title == required_title:
            cell.fill = required_fill

    widths = [15, 12, 50, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord("A") + i - 1)].width = w


def build_evaluation_set_excel_template_bytes(language: str = "zh") -> bytes:
    """Build a downloadable XLSX template.

    Layout:

      Row 1 – instruction / description row
      Row 2 – column headers (Chinese or English per ``language``)
      Row 3+ – example data (multi-turn sessions)
    """
    headers = TEMPLATE_HEADERS.get(language, TEMPLATE_HEADERS["zh"])
    instructions = INSTRUCTIONS.get(language, INSTRUCTIONS["zh"])
    example_rows = EXAMPLE_ROWS.get(language, EXAMPLE_ROWS["zh"])

    wb = Workbook()
    ws = wb.active
    ws.title = "evaluation_cases"

    ws.append(instructions)
    ws.append(headers)
    ws.freeze_panes = "A3"
    _apply_template_styles(ws, language)

    for row in example_rows:
        ws.append(row)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_evaluation_set_export_bytes(
    set_name: str, cases: list[dict[str, Any]]
) -> bytes:
    """Build an XLSX file containing all cases of an evaluation set.

    Produces the same column layout as the zh import template so the file
    can be round-tripped via the upload endpoint.
    """
    headers = TEMPLATE_HEADERS["zh"]
    instructions = INSTRUCTIONS["zh"]

    wb = Workbook()
    ws = wb.active
    ws.title = "evaluation_cases"

    ws.append(instructions)
    ws.append(headers)
    ws.freeze_panes = "A3"
    _apply_template_styles(ws, "zh")

    for case in cases:
        inputs = case.get("inputs") or {}
        label = case.get("label") or {}
        session_id = (
            inputs.get(FIELD_SESSION_ID) or case.get(FIELD_SESSION_ID) or ""
        )
        request_id = (
            inputs.get(FIELD_REQUEST_ID)
            or inputs.get("turn_order")
            or case.get("turn_order")
            or ""
        )
        query = inputs.get(FIELD_QUERY, "")
        answer = label.get(FIELD_ANSWER, "")
        ws.append([session_id, request_id, query, answer])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
